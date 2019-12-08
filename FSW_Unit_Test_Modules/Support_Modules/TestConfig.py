# Anthony Tang anthony.tang@nokia.com
# https://gitlabe2.ext.net.nokia.com/anttang/FSW_Automation.git
# Python 3.7.0

#       standard library imports
import sys
sys.path.insert(0, '')
import os
import tkinter as tk
#       third party imports
import openpyxl as pyxl
#       local imports
from Support_Modules import SpecPane as scpn


class TestConfig():

    testbench_configuration = {}

    def __init__(self, type = "Undefined"):

        self.start_filename = ".\\Resources\\testbench.xlsx"
        self.testbench_configuration = self.parse_config_resource(
                                                    self.start_filename)
        #make a drop down button of all the sheetnames
        self.type = type


    def parse_config_resource(self, filename):
        wb = pyxl.load_workbook(filename)
        ws = wb.active

        read_dic = {}
        #local scaffold instance of the testbench configuration dictionary

        #TODO: use excel to scale the dictionary.
        for index in range(1, 20):
            field_key = "A"+str(index)
            data_key = "B"+str(index)

            field_cell = ws[field_key]
            data_cell = ws[data_key]

            #insert into dictionary
            read_dic[(field_cell.value)] = (data_cell.value)
        print(read_dic)

        return read_dic


    def view_config(self):

        #open popup window displaying toplevel info in specpane format
        window = tk.Toplevel()
        window.title("Testbench Configuration")
        frame = scpn.SpecPane(window, "Testbench Config",
                                    self.testbench_configuration)
        frame.pack()
        edit_btn = tk.Button(window, text = "Edit Config", bg = "yellow",
                                command = lambda: self.open_edit_file() )
        edit_btn.pack()


    def open_edit_file(self):
        #open the config file with excel and maybe it'll just be fine like that
        os.startfile(self.filename)


    def read_default_vals(self):

        if self.type != "Undefined":
            #determine type and load default file.
            if(self.type == 'OBUE'):
                default_file_name = ".\\Resources\\obue_default_vals.xlsx"
            elif(self.type == 'ACLR'):
                default_file_name = ".\\Resources\\aclr_default_vals.xlsx"
            elif(self.type == '5GNR'):
                default_file_name  = ".\\Resources\\5GNR_default_vals.xlsx"

            default_vals = self.parse_config_resource(default_file_name)
            #print(default_vals)

            return default_vals


if __name__ == "__main__":

    TestConfig(type = 'OBUE')
